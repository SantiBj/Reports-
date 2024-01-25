import openpyxl
from ..share.printerDataInSheet import printerDataInSheet
from ..share.calculatedTotals import total
from django.http import HttpResponse
from .fromDictToTuple import formatedFromDictToTuple
from .printerTotal import printerTotal
from ..share.calculatedTotals import calculateTotalWithDataNegative, bookNegativeAndCalculationTotals
from .printers import printerTotalNegatives
from .booksNegToTuple import booksNegToTuple
from .resaltData import resaltDta
from ..share.infDataNegative import createChartText

def defineWorkBook(workBook):
    book = openpyxl.load_workbook(workBook)
    return {
        "book": book,
        "sheet": book.worksheets[0],
    }

def excelCoeditions(dataEditions, cutNumber, hasSap, codCli=False):
    book = None
    sheet = None

    startRow = 5
    startCol = 1

    if codCli:
        recordsCodCli = []
        recordsWithoutDevolutions = []
        devolutions = {}
        firstKey = list(dataEditions.keys())[0]

        if hasSap:
            if dataEditions[firstKey][0]["MONEDA"] == "PESOS":
                defineBook = defineWorkBook("coeSapCOP.xlsx")
                book = defineBook["book"]
                sheet = defineBook["sheet"]
            else:
                defineBook =  defineWorkBook("coeSap.xlsx")
                book = defineBook["book"]
                sheet = defineBook["sheet"]
        else:
            if dataEditions[firstKey][0]["MONEDA"] == "PESOS":
                defineBook = defineWorkBook("coedicionesCOP.xlsx")
                book = defineBook["book"]
                sheet = defineBook["sheet"]
            else:
                defineBook = defineWorkBook("coediciones.xlsx")
                book = defineBook["book"]
                sheet = defineBook["sheet"]

        sheet.cell(row=2, column=2, value=f"Elaborado: {dataEditions[firstKey][0]['ELABORADO']}")
        sheet.cell(row=3, column=2,
                   value=f"Periodo: {dataEditions[firstKey][0]['FECHA']}      Moneda:{dataEditions[firstKey][0]['MONEDA']}")
        

        for nCutCodCli in dataEditions.keys():
            # pinta la cabecera                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                         
            sheet.cell(row=startRow, column=1,
                       value=f"n° corte {nCutCodCli} dependencia => {dataEditions[nCutCodCli][0]['COEDITOR']}")
            startRow += 1
            booksCoeditorWithoutDev = []
            elements = 0
            for recordCodCli in dataEditions[nCutCodCli]:
                recordsCodCli.append(recordCodCli)
                if recordCodCli["CANTIDAD"] >= 0:
                    elements += 1
                    recordsWithoutDevolutions.append(recordCodCli)
                    booksCoeditorWithoutDev.append(
                        formatedFromDictToTuple(recordCodCli, hasSap))
                else :
                    if nCutCodCli in devolutions:
                        oldData = devolutions[nCutCodCli]
                        devolutions[nCutCodCli] = [
                            *oldData,
                            recordCodCli
                        ]
                    else:
                        devolutions[nCutCodCli] = [recordCodCli]

            printerDataInSheet(booksCoeditorWithoutDev, startRow, startCol, sheet)
            startRow = startRow + elements + 1

        #total libros sin devoluciones
        totalWithoutDevolutions = total(recordsWithoutDevolutions)
        printerTotal(sheet, totalWithoutDevolutions, hasSap)
        endRow = sheet.max_row
        resaltDta(sheet,hasSap,endRow)

        #lista de devoluciones
        if len(devolutions.keys()) > 0:
            endRow = sheet.max_row+3
            for numFaculty in devolutions.keys():
                sheet.cell(row=startRow+1, column=1,
                       value=f"n° corte {numFaculty} dependencia => {devolutions[numFaculty][0]['COEDITOR']}")
                printerDataInSheet(booksNegToTuple(
                    devolutions[numFaculty], hasSap,True), endRow, 1, sheet)
                endRow = sheet.max_row+1

            #total devoluciones
            totalsDevolutions = bookNegativeAndCalculationTotals(recordsCodCli)
            resaltDta(sheet,hasSap,endRow)
            printerTotalNegatives(totalsDevolutions,sheet,endRow,hasSap,"Devoluciones : ")
            endRow += 3


            #total libros y devoluciones
            totalEnd = calculateTotalWithDataNegative(totalWithoutDevolutions,totalsDevolutions)
            resaltDta(sheet,hasSap,endRow)
            printerTotalNegatives(totalEnd,sheet,endRow,hasSap,"Total : ")

            createChartText(sheet,hasSap,coe=True)

        dateOfFile = dataEditions[firstKey][0]['FECHA'].split('-')

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={dataEditions[nCutCodCli][0]["COEDITOR"][4:30]}{dateOfFile[4]}_{dateOfFile[3]}_{cutNumber}.xlsx'
        book.save(response)
        return response

    else:
        if hasSap:
            if dataEditions[0]["MONEDA"] == "PESOS":
                defineBook = defineWorkBook("coeSapCOP.xlsx")
                book = defineBook["book"]
                sheet = defineBook["sheet"]
            else:
                defineBook =  defineWorkBook("oeSap.xlsx")
                book = defineBook["book"]
                sheet = defineBook["sheet"]
        else:
            if dataEditions[0]["MONEDA"] == "PESOS":
                defineBook = defineWorkBook("coedicionesCOP.xlsx")
                book = defineBook["book"]
                sheet = defineBook["sheet"]
            else:
                defineBook = defineWorkBook("coediciones.xlsx")
                book = defineBook["book"]
                sheet = defineBook["sheet"]

        sheet.cell(row=2, column=2, value=f"Coeditor: {dataEditions[0]['COEDITOR']}          Elaborado: {dataEditions[0]['ELABORADO']}")
        sheet.cell(
            row=3, column=2, value=f"Periodo: {dataEditions[0]['FECHA']}     N° corte: {cutNumber}     Moneda: {dataEditions[0]['MONEDA']}")
        
        devolutions = []
        booksWithoutDevolutions = []
        booksWithoutDevDict = []

        for record in dataEditions:
            if record["CANTIDAD"] >= 0:
                booksWithoutDevDict.append(record)
                booksWithoutDevolutions.append(formatedFromDictToTuple(record, hasSap))
            else :
                devolutions.append(record)

        # libros sin devoluciones
        printerDataInSheet(booksWithoutDevolutions, startRow, startCol, sheet)

        # total libros sin devoluciones
        totalWithoutDev = total(booksWithoutDevDict)
        printerTotal(sheet, totalWithoutDev, hasSap)
        endRow = sheet.max_row
        resaltDta(sheet,hasSap,endRow)

        # lista de devoluciones
        if len(devolutions) > 0:
            totalsNegatives = bookNegativeAndCalculationTotals(dataEditions)
            endRow = sheet.max_row + 2
            printerDataInSheet(booksNegToTuple(
                totalsNegatives, hasSap), endRow, 1, sheet)
            endRow = endRow = sheet.max_row+1

            #Total devoluciones
            totalsDevolutions = bookNegativeAndCalculationTotals(dataEditions)
            resaltDta(sheet,hasSap,endRow)
            printerTotalNegatives(totalsDevolutions,sheet,endRow,hasSap,"Devoluciones : ")

            endRow += 3

            #Total neto
            totalEnd = calculateTotalWithDataNegative(totalWithoutDev,totalsDevolutions)
            resaltDta(sheet,hasSap,endRow)
            printerTotalNegatives(totalEnd,sheet,endRow,hasSap,"Total : ")

            createChartText(sheet,hasSap,coe=True)

        dateOfFile = dataEditions[0]['FECHA'].split('-')

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={dataEditions[0]["COEDITOR"][4:30]}{dateOfFile[4]}_{dateOfFile[3]}_{cutNumber}.xlsx'
        book.save(response)
        return response
