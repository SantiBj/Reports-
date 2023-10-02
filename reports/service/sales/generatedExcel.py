import openpyxl
from django.http import HttpResponse
from ..share.calculatedTotals import printerTotalBooksAndDevolutions
from datetime import datetime
from .formatedRecordForDictToTuple import formatedRecordForDictToTuple
from ..share.printerDataInSheet import printerDataInSheet
from .bookWithoutDevolutions import bookWithoutDevolutions

def generatedExcel(dataDict, cutNumber, hasSap):
    book = None
    if hasSap:
        if dataDict[0]["MONEDA"] != "PESOS":
            book = openpyxl.load_workbook("plantilla-UEX.xlsx")
        else:
            book = openpyxl.load_workbook("plantilla-UEX_COP.xlsx")
    else:
        if dataDict[0]["MONEDA"] != "PESOS":
            book = openpyxl.load_workbook("plantilla-reporte-ventas.xlsx")
        else:
            book = openpyxl.load_workbook("plantilla-reporte-ventas_COP.xlsx")

    sheet = book.worksheets[0]

    startRow = 5
    startCol = 1

    # añadiendo cabecera
    sheet.cell(row=2, column=2,
               value=f"Proveedor: {dataDict[0]['PROVEEDOR']}           Elaborado:{dataDict[0]['ELABORADO']}")
    sheet.cell(row=3, column=2,
               value=f"Periodo: {dataDict[0]['FECHA']}     N° corte: {cutNumber}        Moneda: {dataDict[0]['MONEDA']}")

    # extrayendo los libros sin las devoluciones
    books = bookWithoutDevolutions(dataDict,hasSap)
    printerDataInSheet(books["booksTup"], startRow, startCol, sheet)

    # añadiendo los totales al excel tanto libros como devoluciones
    printerTotalBooksAndDevolutions(
        books["booksDict"], dataDict, sheet, hasSap)

    date = dataDict[0]["FECHA"].split('-')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={dataDict[0]["PROVEEDOR"][:3]}{date[4]}_{date[3]}-{cutNumber}.xlsx'
    book.save(response)
    return response



