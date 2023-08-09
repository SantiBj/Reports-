import openpyxl
from django.http import HttpResponse
from .calculatedTotals import calculatedTotalsBySupplier
from datetime import datetime
from .formatedRecordForDictToTuple import formatedRecordForDictToTuple



def generatedExcel(dataDict, cutNumber):
    isUEX = True if "SAP" in dataDict[0] else False
    # si es UEX usar otra plantilla y otro formateador
    dataInTuples = []
    for record in dataDict:
        dataInTuples.append(formatedRecordForDictToTuple(record,isUEX))

    book = openpyxl.load_workbook("plantilla-UEX.xlsx" if isUEX else "plantilla-reporte-ventas.xlsx")
    sheet = book.worksheets[0]

    startRow = 5
    startCol = 1

    # add name supplier in excel and n° cut
    sheet.cell(row=2, column=2, value=f"Proveedor: {dataDict[0]['PROVEEDOR']}")
    sheet.cell(row=3, column=2,
               value=f"Periodo: {datetime.now().date()}     N° corte: {cutNumber}")

    for rowIndex, rowData in enumerate(dataInTuples, start=startRow):
        for colIndex, value in enumerate(rowData, start=startCol):
            sheet.cell(row=rowIndex, column=colIndex, value=value)

    # calculo de total cantidad,bruto,neto
    calculatedTotalsBySupplier(dataDict, sheet,isUEX)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={dataDict[0]["PROVEEDOR"][:3]}.xlsx'
    book.save(response)
    return response
